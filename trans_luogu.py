from openpyxl import load_workbook
import pandas
import json

# 加载Excel文件
wb = load_workbook('team.xlsx',data_only=True)

# 选择工作表
ws = wb['Sheet1']

ans={}

flag=0
# 遍历行和列来读取数据
for row in ws.iter_rows(max_row=ws.max_row,values_only=True):
    flag+=1

    if flag <= 2: # 排除表格中第一、二行的非队伍信息行
        continue
    if row[0]==None:#读到空行结束
        break

    ansItem={}
    ansItem['team_id']=row[4]
    ansItem['name']=row[15]#对应洛谷昵称
    ansItem['team_name']=row[1]#队伍名称
    ansItem['organization']=row[2]#所属学校
    ansItem['location']=row[14]#队伍位置
    if row[1][0]=='*':
        ansItem['official']=False
        ansItem['unofficial']=True
    else:
        ansItem['official']=True
        ansItem['unofficial']=False
    if row[9]=="N":
        ansItem['girl']=False
    else:
        ansItem['girl']=True
    ansItem['group'] = []
    ansItem['coach']=row[11]
    ansItem['members']=[]
    ansItem['members'].append(row[5])
    ansItem['members'].append(row[6])
    ansItem['members'].append(row[7])
    ans[ansItem['team_id']]=ansItem
    print("第%d行已生成json\n" %flag)

# 指定要保存的文件名
filename = "./data/team.json"

# 打开文件并写入JSON数据
with open(filename, 'w', encoding='utf-8') as f:
    json.dump(ans, f, ensure_ascii=False, indent=4)

print(f"数据已成功保存到 {filename}")