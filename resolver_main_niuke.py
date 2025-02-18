import json
import sys
from time import mktime
import resolver_functions
import resolver_utils
from datetime import datetime, timezone, timedelta
from openpyxl import load_workbook
import gzip
import urllib.request
import pandas
import time

def convert_ms_to_time(timestamp):
    # 将毫秒转换为秒，并分解为整秒和毫秒
    total_seconds = timestamp / 1000
    seconds = int(total_seconds)
    milliseconds = int((total_seconds - seconds) * 1000)

    # 分解秒为分钟和秒
    minutes = seconds // 60
    seconds %= 60
    hours = minutes // 60
    minutes = minutes % 60

    # 格式化输出
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}.{milliseconds:03d}"

def convert_ms_to_isoformat(ms, tz_hours=8):
    # 创建timezone对象，这里假设时区为东八区
    tz = timezone(timedelta(hours=tz_hours))

    # 将毫秒时间戳转换为datetime对象，并设置时区
    dt = datetime.fromtimestamp(ms / 1000, tz=tz)

    # 格式化为ISO 8601格式，注意Python默认会根据时区调整时间，因此不需要手动处理偏移量
    iso_time = dt.isoformat(timespec='milliseconds')

    return iso_time

#给题目编号
pid={"万物起源":1,"Fever":2,"一闪一闪亮晶晶":3,"宇宙跳跃":4,"收手吧，Bob":5,"薯条日！":6,"隐秘角落":7,"&quot;原&quot;题":8,"献给你的花束":9,"《我不是烟神》":10}
#-----------------------------------------------写入run.json文件
# 加载Excel文件
wb = load_workbook('sub.xlsx')
# 选择工作表
ws = wb['Sheet1']

flag=0
ans=[]#用列表存储提交信息
# 遍历行和列来读取数据
for row in ws.iter_rows(max_row=ws.max_row,values_only=True):
    if flag < 1: # 排除表格中第一行的非提交信息行
        flag += 1
        continue
    if row[0]==None:#读到空行结束
        break
    ansItem={}#用字典存储一个队伍的信息
    ansItem['team_id']=row[1]#队伍ID
    ansItem['problem_id']=pid[row[2]]#题目ID
    tim=datetime.strptime(row[5],"%Y-%m-%d %H:%M:%S")
    tim_tp=tim.timetuple()
    ansItem['timestamp']=int(time.mktime(tim_tp))#计算罚时
    ansItem['language']=row[3]
    ansItem['submission_id']=row[0]
    ansItem['status']=row[4]
    ans.append(ansItem)
# 指定要保存的文件名
filename = "./data/run.json"
# 打开文件并写入JSON数据
with open(filename, 'w', encoding='utf-8') as f:
    json.dump(ans, f, ensure_ascii=False, indent=4)
#--------------------------------------------------------
with open("./data/resolver.json", 'w', encoding='utf-8') as ff:
    # 读入配置文件
    config=""
    with open('./config/config.json', 'r', encoding='utf-8') as f:
        config = json.load(f)
    # 读入配置文件
    cfg=""
    with open('./config/cfg.json', 'r', encoding='utf-8') as f:
        cfg = json.load(f)


    # pta状态和resolver状态的对应
    resultDict={}
    resultDict['答案正确']='AC'
    resultDict['答案错误']='WA'
    resultDict['运行超时']='TLE'
    resultDict['返回非零']='NZEC'
    resultDict['编译错误']='CE'
    resultDict['格式错误']='PE'
    resultDict['段错误']='SF'
    resultDict['内存超限']='MLE'
    resultDict['运行错误']='RE'
    resultDict['输出超限']='OLE'

    # 添加state
    state={}
    state['type']="state"
    stateData={}
    state['data']=stateData
    state['token']='cdT0'

    ff.write(json.dumps(state) + '\n')


    # 添加contest
    contest={}
    contest['type']="contest"
    contest['id']=cfg['problemId']
    contestData={}
    contestData['id']=cfg['problemId']
    contestData['name']=config['contest_name']
    #print("比赛昵称："+contestData['name'])
    contestData['formal_name']=config['contest_name']
    contestData['start_time']=resolver_utils.transTimestampToTime(config['start_time'])
    contestData['duration']=resolver_utils.calculateTimeDiff(config['start_time'],config['end_time'])
    contestData['scoreboard_freeze_duration']=resolver_utils.calculateTimeDiff(0,config['frozen_time'])
    contestData['penalty_time']=int(config['penalty']/60)
    contestData['scoreboard_type']="pass-fail"
    contest['data']=contestData
    contest['token']="cdT1"

    ff.write(json.dumps(contest) + '\n')

    # 添加judgement
    i=2
    contestJudgement=[]
    with open('./config/contest_judgement.json', 'r', encoding="utf-8") as f:
        contestJudgement = json.load(f)

    judgements=[]
    for v in contestJudgement:
        judgement={}
        judgement['type']='judgement-types'
        judgement['id']=v['acronym']
        judgementData={}
        judgementData['id']=v['acronym']
        judgementData['name']=v['name']
        judgementData['penalty']=v['penalty']
        judgementData['solved']=v['solved']
        judgement['data']=judgementData
        judgement['token']='cdT'+str(i)
        judgements.append(judgement)
        i=i+1

    for v in judgements:
        ff.write(json.dumps(v) + '\n')


    # 添加language
    contestLanguage=[]
    with open('./config/contest_language.json', 'r', encoding="utf-8") as f:
        contestLanguage = json.load(f)

    languages=[]
    for v in contestLanguage:
        language={}
        language['type']='languages'
        language['id']=v['id']
        languageData={}
        languageData['id']=v['id']
        languageData['name']=v['name']
        languageData['entry_point_required']=False
        language['data']=languageData
        language['token']='cdT'+str(i)
        languages.append(language)
        i=i+1

    for v in languages:
        ff.write(json.dumps(v) + '\n')

    # 添加problems
    contestProblem=[]
    with open('./config/contest_problem.json', 'r', encoding="utf-8") as f:
        contestProblem = json.load(f)

    problems=[]
    for v in contestProblem:
        problem={}
        problem['type']='problems'
        problem['id']=v['id']
        problemData={}
        problemData['id']=v['id']
        problemData['label']=v['letter']
        problemData['name']=v['name']
        problemData['uuid']=v['id']
        problemData['ordinal']=v['id']
        problemData['color']=v['color']
        problemData['rgb']=v['rgb']
        problem['data']=problemData
        problem['token']='cdT'+str(i)
        problems.append(problem)
        i=i+1

    for v in problems:
        ff.write(json.dumps(v) + '\n')

    # 添加groups，可以用来设置区域奖项
    # 但是我们没有，我加这个纯凑数，怕少这一项报错
    # 想加的话可以改进一下，注意下边的team项中的group_id也需要修改
    group = {}
    group['type'] = "groups"
    group['id'] = '1'
    groupData = {}
    groupData['id'] = '1'
    groupData['name'] = "正式队伍"
    group['data'] = groupData
    group['token'] = 'cdT' + str(i)
    i = i + 1

    ff.write(json.dumps(group) + '\n')
    group = {}
    group['type'] = "groups"
    group['id'] = '2'
    groupData = {}
    groupData['id'] = '2'
    groupData['name'] = "打星队伍"
    group['data'] = groupData
    group['token'] = 'cdT' + str(i)
    i = i + 1
    ff.write(json.dumps(group) + '\n')

    # 添加队伍
    contestTeam={}
    with open('./data/team.json', 'r', encoding="utf-8") as f:
        contestTeam = json.load(f)
    teams=[]
    for key,value in contestTeam.items():
        team={}
        team['type']='teams'
        team['id']=value['team_id']
        teamData={}
        teamData['id']=value['team_id']
        # if value['girl']==True:#滚榜展示的队伍名称
        #     teamData['name']=value['team_name']+"(女生队)"
        # else:
        teamData['name']=value['team_name']
        if value['official']== True:
             teamData['group_ids']=['1']
        else:teamData['group_ids']=['2']
        teamData['university']=value['organization']
        team['data']=teamData
        team['token']="cdT"+str(i)
        teams.append(team)
        i=i+1

    for v in teams:
        ff.write(json.dumps(v) + '\n')

    # 添加提交记录
    contestRun=[]
    with open('./data/run.json', 'r', encoding="utf-8") as f:
        contestRun = json.load(f)
    with open('./config/config.json','r',encoding="utf-8") as pot:
        cpt=json.load(pot)
    submissions=[]
    judgements=[]
    num=0#记录第几条cdT
    for j in range(0,len(contestRun)):
        v=contestRun[j]

        submission={}
        submission['type']="submissions"
        submission['id']=v['submission_id']
        submissionData={}
        submissionData['id']=v['submission_id']
        submissionData['problem_id']=str(v['problem_id'])
        submissionData['team_id']=v['team_id']
        submissionData['language_id']=v['language']
        submissionData['files']=[]
        submissionData['contest_time']=convert_ms_to_time((v['timestamp']-cpt["start_time"])*1000)
        submissionData['time']=convert_ms_to_isoformat(config['start_time']+v['timestamp'])
        submission['data']=submissionData
        submission['token']='cdT'+str(i)
        submissions.append(submission)
        i=i+1

        judgement={}
        judgement['type']="judgements"
        judgement['id']=v['submission_id']
        judgementData={}
        judgementData['id']=v['submission_id']
        judgementData['submission_id']=v['submission_id']
        judgementData['judgement_type_id']=resultDict[v['status']]

        judgementData['start_contest_time']=convert_ms_to_time(v['timestamp'])
        judgementData['start_time']=convert_ms_to_isoformat(config['start_time']+v['timestamp'])
        if judgementData['judgement_type_id']=='AC':
            judgementData['score']=300.0
        else:
            judgementData['score']=0.0
        judgement['data']=judgementData
        judgement['token']='cdT'+str(i)
        judgements.append(judgement)
        i=i+1
        num=i

    for j in range(0,len(submissions)):
        ff.write(json.dumps(submissions[j]) + '\n')
        ff.write(json.dumps(judgements[j]) + '\n')


    # 添加获奖设置
    # 最好不写这部分，使用award手动设置奖项，更加精细不会出错
    # contestAwards={}
    # with open('./config/contest_awards.json', 'r', encoding="utf-8") as f:
    #     contestAwards = json.load(f)

    # for v in contestAwards:
    #     ff.write(json.dumps(v) + '\n')

    #添加finalized 以下部分必须有且不修改最好
    state = {}
    state['type'] = "state"
    stateData = {}
    stateData['started'] = resolver_utils.transTimestampToTime(config['start_time'])
    stateData['frozen'] = resolver_utils.transTimestampToTime(config['end_time'] - config['frozen_time'])
    stateData['thawed']=resolver_utils.transTimestampToTime(config['end_time'] - config['frozen_time'])
    stateData['finalized']=resolver_utils.transTimestampToTime(config['end_time'])
    stateData['end_of_updates'] = resolver_utils.transTimestampToTime(config['end_time'])
    state['data'] = stateData
    state['token'] = 'cdT'+str(num)
    num=num+1
    ff.write(json.dumps(state) + '\n')

    state = {}
    state['type'] = "state"
    stateData = {}
    stateData['started'] = resolver_utils.transTimestampToTime(config['start_time'])
    stateData['ended']=resolver_utils.transTimestampToTime(config['end_time'])
    stateData['frozen'] = resolver_utils.transTimestampToTime(config['end_time'] - config['frozen_time'])
    stateData['thawed'] = resolver_utils.transTimestampToTime(config['end_time'] - config['frozen_time'])
    stateData['finalized'] = resolver_utils.transTimestampToTime(config['end_time'])
    stateData['end_of_updates'] = resolver_utils.transTimestampToTime(config['end_time'])
    state['data'] = stateData
    state['token'] = 'cdT' + str(num)

    ff.write(json.dumps(state) + '\n')

print("文件生成成功")